"""
report_generator.py
--------------------
Generates downloadable reports (PDF, Excel, CSV) summarizing candidate
screening results. Kept separate from UI code so reports can be
produced/tested independently of Streamlit.
"""

import io
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
)

from ai.ranking import candidates_to_dataframe


def _deduplicate_candidates(candidates: List[dict]) -> List[dict]:
    """Remove duplicate candidates by file_hash, keeping the first occurrence."""
    seen = set()
    unique = []
    for c in candidates:
        h = c.get("file_hash", "")
        if h and h in seen:
            continue
        if h:
            seen.add(h)
        unique.append(c)
    return unique


def generate_csv(candidates: List[dict]) -> bytes:
    # Deduplicate by file_hash
    unique = _deduplicate_candidates(candidates)
    df = candidates_to_dataframe(unique)
    
    # Ensure all candidates have required fields, including rejected ones
    if len(df) == 0:
        # Return empty CSV with headers if no candidates
        df = pd.DataFrame(columns=["Rank", "Candidate", "Email", "Phone", "Experience (yrs)", 
                                   "Education", "Match Score", "Status", "Recommendation", 
                                   "Missing Skills", "Uploaded"])
    
    return df.to_csv(index=False).encode("utf-8")


def generate_excel(candidates: List[dict], job: dict) -> bytes:
    # Deduplicate by file_hash
    unique = _deduplicate_candidates(candidates)
    df = candidates_to_dataframe(unique)

    wb = Workbook()
    ws = wb.active
    ws.title = "Candidate Ranking"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in df.itertuples(index=False):
        ws.append(list(row))

    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col]]) if len(df) else len(col)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 45)

    # Job description sheet
    ws2 = wb.create_sheet("Job Description")
    ws2.append(["Field", "Value"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for key in ["title", "department", "location", "experience_years", "education", "skills"]:
        val = job.get(key, "")
        if isinstance(val, list):
            val = ", ".join(val)
        ws2.append([key.replace("_", " ").title(), val])
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 50

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_pdf(candidates: List[dict], job: dict) -> bytes:
    # Deduplicate by file_hash
    candidates = _deduplicate_candidates(candidates)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleStyle", parent=styles["Title"], textColor=colors.HexColor("#1F4E79"),
        fontSize=20, spaceAfter=6,
    )
    heading_style = ParagraphStyle(
        "HeadingStyle", parent=styles["Heading2"], textColor=colors.HexColor("#1F4E79"),
        spaceBefore=14, spaceAfter=8,
    )
    normal = styles["Normal"]
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=9, leading=12)

    elements = []
    elements.append(Paragraph("AI Resume Screening Report", title_style))
    elements.append(Paragraph(
        f"Job Title: <b>{job.get('title', 'N/A')}</b> | Department: {job.get('department', 'N/A')} "
        f"| Location: {job.get('location', 'N/A')}", normal
    ))
    elements.append(Paragraph(
        f"Required Experience: {job.get('experience_years', 0)} yrs | "
        f"Required Education: {job.get('education', 'N/A')}", normal
    ))
    elements.append(Paragraph(
        f"Required Skills: {', '.join(job.get('skills', [])) or 'N/A'}", normal
    ))
    elements.append(Spacer(1, 12))

    # Ranking table
    elements.append(Paragraph("Candidate Ranking Summary", heading_style))
    table_data = [["Rank", "Candidate", "Score", "Status", "Missing Skills"]]
    for c in candidates:
        scores = c.get("scores", {})
        missing = ", ".join(scores.get("missing_skills", [])[:4]) or "None"
        table_data.append([
            str(c.get("rank", "-")),
            c.get("name", "Unknown"),
            f"{scores.get('overall_score', 0)}%",
            scores.get("status", "-"),
            missing,
        ])

    tbl = Table(table_data, colWidths=[1.5 * cm, 4 * cm, 2 * cm, 2.7 * cm, 6.5 * cm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(tbl)
    elements.append(PageBreak())

    # Detailed candidate profiles
    elements.append(Paragraph("Detailed Candidate Profiles", heading_style))
    for c in candidates:
        scores = c.get("scores", {})
        if not scores:
            continue
            
        elements.append(Paragraph(
            f"#{c.get('rank', '-')} - {c.get('name', 'Unknown')} "
            f"({scores.get('overall_score', 0)}% - {scores.get('status', '-')})",
            styles["Heading3"],
        ))
        elements.append(Paragraph(
            f"Email: {c.get('email', 'Not Available')} | Phone: {c.get('phone', 'Not Available')}", small
        ))
        
        # Handle missing education/location gracefully
        education = c.get('education', [])
        education_text = ', '.join(education) if education else 'Not Available'
        location = scores.get('location', 'Not Found') or 'Not Found'
        experience = c.get('experience_years', 0)
        
        elements.append(Paragraph(
            f"Experience: {experience} yrs | "
            f"Education: {education_text} | "
            f"Location: {location}", small
        ))
        
        skills = c.get('skills', [])
        skills_text = ', '.join(skills) if skills else 'Not Available'
        elements.append(Paragraph(
            f"Skills: {skills_text}", small
        ))
        
        missing_skills = scores.get('missing_skills', [])
        missing_text = ', '.join(missing_skills[:10]) if missing_skills else 'None'
        elements.append(Paragraph(
            f"Missing Skills: {missing_text}", small
        ))
        
        # Location explanation
        location_explanation = scores.get('location_explanation', '')
        if location_explanation:
            elements.append(Paragraph(f"Location Match: {location_explanation}", small))
        
        if scores.get("strengths"):
            elements.append(Paragraph("Strengths:", small))
            for s in scores["strengths"][:10]:  # Limit to top 10
                elements.append(Paragraph(f"&#10003; {s}", small))
        if scores.get("weaknesses"):
            elements.append(Paragraph("Weaknesses:", small))
            for w in scores["weaknesses"][:10]:  # Limit to top 10
                elements.append(Paragraph(f"&#10007; {w}", small))
        elements.append(Paragraph(f"<b>Recommendation:</b> {scores.get('summary', 'No summary available.')}", small))
        elements.append(Spacer(1, 10))

    doc.build(elements)
    return buffer.getvalue()
